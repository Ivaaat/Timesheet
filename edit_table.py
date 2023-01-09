import datetime
import calendar, locale
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
                        PatternFill, Border, Side, 
                        Alignment, Font
                        )
import win32com.client
import pythoncom
from PIL import ImageGrab
import psutil
import time
from ctypes import windll


def except_perm(filename):
    for process in psutil.process_iter():
        if process.name() == 'EXCEL.EXE':
            for open_file in process.open_files():
                if filename in open_file.path:
                    process.terminate()
                    break

def edit_timesheet(filename, range_work, employee_name):
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        line_shift = 3
        wb = load_workbook(filename = filename)
        sheet_ranges = wb[employee_name]
        for range_list in range_work:
            for i in range(int(range_list[0]),int(range_list[1]) + 1):
                sheet_ranges[f'B{i + line_shift}'] = range_list[2]
                sheet_ranges[f'C{i + line_shift}'] = range_list[3]
                sheet_ranges[f'D{i + line_shift}'] = range_list[4]
                if len(range_list[4]) < len("Комментарии"):
                    sheet_ranges.column_dimensions['D'].width = 15
                else:
                    sheet_ranges.column_dimensions['D'].width = len(range_list[4]) * 1.15
        wb.save(filename)



def default_timesheet(place="Работа в офисе",employee_name="Конихин Иван Владимирович"):
#def default_timesheet(place="Работа в офисе", employee_name="Иванов Иван Иваныч"):
    weekends_and_holidays2022 = {'Январь' : [1,2,3,4,5,6,7,8,9],
                                'Февраль': [23],
                                'Март': [6,8],
                                "Май": [1,2,3,7,8,9,10],
                                'Июнь': [11,12,13],
                                'Ноябрь':[4,5,6]}

    weekends_and_holidays2023 = {'Январь' : [1,2,3,4,5,6,7,8],
                                'Февраль': [23,24,25,26],
                                'Март': [8],
                                'Апрель': [29,30],
                                "Май": [6,7,8,9],
                                'Июнь': [10,11,12],
                                'Ноябрь':[4,5,6]}

    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
    now = datetime.datetime.now()
    year = now.year
    month = now.month
    line_shift = 3
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    green_color = '92d050'
    #red_color = 'ff0000'
    red_color = 'ff4343'
    
    wb = Workbook()
    dict_range = {'A1':"ФИО:",
                'A2':"Месяц:",
                'A3':"Дата",
                'B1':employee_name,
                'B2':f"{calendar.month_name[month]} {str(year)}",
                'B3': "Описание выполненных работ",
                'C1': "",
                'C2': "",
                'C3': "Время работы, ч",
                'D1': "",
                'D2': "",
                'D3': "Комментарии",
    }

    filename = f'Табель {calendar.month_name[month]  } {year}.xlsx'
    try:
        wb = load_workbook(filename = filename)
    except IOError:
        #ws = wb.create_sheet(employee_name, 0)
        ws = wb.active
        ws.title = employee_name
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        for range, descr in dict_range.items():
            ws[range].font = Font(color="000000", bold=True)
            ws[range].border = Border(top=medium, bottom=medium, left=medium, right=medium)
            ws[range].fill = PatternFill('solid', fgColor=green_color)
            if not range.startswith("A"):
                ws[range].alignment = Alignment(horizontal='center')
            ws[range] = descr
    if employee_name in wb.sheetnames:
        ws = wb[employee_name]   
    else:
        ws = wb.create_sheet(employee_name)
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        for range, descr in dict_range.items():
            ws[range].font = Font(color="000000", bold=True)
            ws[range].border = Border(top=medium, bottom=medium, left=medium, right=medium)
            ws[range].fill = PatternFill('solid', fgColor=green_color)
            if not range.startswith("A"):
                ws[range].alignment = Alignment(horizontal='center')
            ws[range] = descr
    month_calendar = calendar.monthcalendar(year,month)
    asd = []
    list_name_day = list(calendar.day_abbr) 
    for list_week in month_calendar:
        calendar_month = {}
        for i, day in enumerate(list_week):
            calendar_month[list_name_day[i]] = day
        asd.append(calendar_month)
    list_range = ['A', 'B', 'C', 'D']
    i = 0
    for dict_week in asd:
        for name_day, num_day in dict_week.items():
            try:
                weekends_and_holidays = weekends_and_holidays2023[calendar.month_name[month]] 
            except KeyError:
                weekends_and_holidays = {}
            if name_day in ['Сб','Вс'] and num_day != 0 or num_day in weekends_and_holidays:
                i += 1
                work_and_travel = ""
                num_work_hour = ""
                for name_range in list_range:
                    if name_range == 'A':
                        ws[f'{name_range}{num_day + line_shift}'].font = Font(color="000000", bold=True)
                        ws[f'{name_range}{num_day + line_shift}'].border = Border(top=thin, bottom=thin, left=thin, right=medium)

                    else:
                        ws[f'{name_range}{num_day + line_shift}'].font = Font(color="000000", bold=False)
                        ws[f'{name_range}{num_day + line_shift}'].border = Border(top=thin, bottom=thin, left=thin, right=thin)
                    ws[f'{name_range}{num_day + line_shift}'].fill = PatternFill('solid', fgColor=red_color)
            elif num_day != 0:
                #if now.day <= i and now.day <= 15:
                i += 1
                if  15 < i and now.day < 15:
                    work_and_travel = ""
                    num_work_hour = ""
                else:
                    work_and_travel = place
                    num_work_hour = 8
                for name_range in list_range:
                    if name_range == 'A':
                        ws[f'{name_range}{num_day + line_shift}'].fill = PatternFill('solid', fgColor=green_color)
                        ws[f'{name_range}{num_day + line_shift}'].font = Font(color="000000", bold=True)
                        ws[f'{name_range}{num_day + line_shift}'].border = Border(top=thin, bottom=thin, left=thin, right=medium)
                        
                    else:
                        ws[f'{name_range}{num_day + line_shift}'].fill = PatternFill('solid', fgColor="ffffff")
                        ws[f'{name_range}{num_day + line_shift}'].font = Font(color="000000", bold=False)
                        ws[f'{name_range}{num_day + line_shift}'].border = Border(top=thin, bottom=thin, left=thin, right=thin)
            else:
                continue
            ws[f'A{num_day + line_shift}'] = name_day + " " + str(num_day)
            ws[f'B{num_day + line_shift}'] = work_and_travel
            ws[f'C{num_day + line_shift}'] = num_work_hour
            ws[f'D{num_day + line_shift}'] = ""
            ws.column_dimensions['D'].width = 15
    #wb.save(filename)
    #return filename
    try:
        wb.save(filename)
        return filename
    except OSError:
         except_perm(filename)
         default_timesheet(employee_name = employee_name)
         wb.save(filename)
         return filename

    

    #thins = Side(border_style="medium", color="000000")
    
#teams = ['Грязнов Юрий Сергеевич', 'Дыдыкин Иван Александрович', "Храмов Роман Андреевич", "Смирнов Александр Юрьевич"]
#for name in teams:
    #default_timesheet(employee_name = name)

def close_workbook(filename):
    excel = win32com.client.Dispatch("Excel.Application", pythoncom.CoInitialize())
    path = get_path(filename)
    wb_path = rf'{path}'
    wb = excel.Workbooks.Open(wb_path)
    wb.Close()

def export_excel_jpeg(filename, name_sheet):
    #o = win32com.client.Dispatch("Excel.Application", pythoncom.CoInitialize())
    o = win32com.client.DispatchEx("Excel.Application", pythoncom.CoInitialize())
    o.Visible = False
    path = get_path(filename)
    wb_path = rf'{path}'
    wb1 = o.Workbooks.Open(wb_path)
    ws = wb1.Worksheets(name_sheet)
    #time.sleep(1)
    try:
        ws.Range("A1:D34").CopyPicture(Format = 2)
    except Exception:
        export_excel_jpeg(filename, name_sheet)
    path_to_jpeg = r'{}'.format(path[:path.find('.')] + '.jpg')
    img = ImageGrab.grabclipboard()
    img.save(get_path(path_to_jpeg) )
    wb1.Close()
    o.CutCopyMode = False
    o.Quit()
    return path_to_jpeg


def get_path(*path):
    root_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(root_path, *path)


